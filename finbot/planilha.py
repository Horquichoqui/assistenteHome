"""Geração da planilha financeira (.xlsx) com openpyxl, no formato do casal:
resumo do mês, gastos fixos/variáveis, faturas de cartão, orçamento por
categoria com teto, investimentos e as abas históricas de sempre."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import stats
from .config import Config
from .db import Database, Gasto

_FORMATO_MOEDA = '"R$" #,##0.00'
_COR_CABECALHO = "1F4E78"
_COR_RENDA = "C6EFCE"
_COR_DESPESA = "FFC7CE"
_COR_DESTAQUE = "FFF2CC"

_MESES = [
    "janeiro", "fevereiro", "março", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
]


def _cabecalho(ws: Worksheet, titulos: list[str], larguras: list[int], linha: int = 1) -> None:
    fill = PatternFill("solid", fgColor=_COR_CABECALHO)
    fonte = Font(bold=True, color="FFFFFF")
    for col, (titulo, largura) in enumerate(zip(titulos, larguras), start=1):
        celula = ws.cell(row=linha, column=col, value=titulo)
        celula.font = fonte
        celula.fill = fill
        celula.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = largura
    ws.freeze_panes = ws.cell(row=linha + 1, column=1).coordinate


def _linha_destaque(ws: Worksheet, linha: int, rotulo: str, valor_centavos: int, cor: str) -> None:
    fill = PatternFill("solid", fgColor=cor)
    c1 = ws.cell(row=linha, column=1, value=rotulo)
    c1.font = Font(bold=True)
    c1.fill = fill
    c2 = ws.cell(row=linha, column=2, value=valor_centavos / 100)
    c2.number_format = _FORMATO_MOEDA
    c2.fill = fill
    c2.font = Font(bold=True)


def _linhas_gastos(
    ws: Worksheet, gastos: list[Gasto], cfg: Config, linha_inicial: int
) -> int:
    """Escreve linhas no layout Responsável/Categoria/Descrição/Data/Parcela/
    Pagamento/Valor/Pago?. Retorna a próxima linha livre."""
    linha = linha_inicial
    for g in gastos:
        ws.cell(row=linha, column=1, value=cfg.nome_pessoa(g.responsavel))
        ws.cell(row=linha, column=2, value=g.categoria)
        ws.cell(row=linha, column=3, value=g.estabelecimento or g.descricao or "")
        ws.cell(row=linha, column=4, value=g.data_compra).number_format = "DD/MM/YYYY"
        ws.cell(row=linha, column=5, value=g.parcela_rotulo or "")
        ws.cell(row=linha, column=6, value=g.cartao_nome or g.forma_pagamento or "")
        ws.cell(row=linha, column=7, value=g.valor).number_format = _FORMATO_MOEDA
        ws.cell(row=linha, column=8, value="Sim" if g.pago else "Não")
        linha += 1
    return linha


def gerar_planilha(db: Database, destino: Path, hoje: date, cfg: Config) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    ini_mes, fim_mes = stats.limites_do_mes(hoje)
    gastos_mes = db.listar_gastos(ini_mes, fim_mes)
    cartoes = db.listar_cartoes()
    rendas = db.listar_rendas()
    investimentos = db.listar_investimentos()
    tetos = db.listar_tetos()

    renda_total = stats.total_renda_centavos(rendas)
    despesas_total = stats.total_centavos(gastos_mes)
    investimentos_total = stats.total_investimentos_centavos(investimentos)
    saldo = renda_total - despesas_total

    # ================= Resumo do Mês =================
    ws = wb.create_sheet("Resumo do Mês")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 16

    titulo = ws.cell(row=1, column=1, value=f"Nosso Resumo de {_MESES[hoje.month - 1]} 💰")
    titulo.font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    _linha_destaque(ws, 3, "RENDA TOTAL", renda_total, _COR_RENDA)
    _linha_destaque(ws, 4, "DESPESAS TOTAIS", despesas_total, _COR_DESPESA)
    _linha_destaque(ws, 5, "INVESTIMENTOS", investimentos_total, _COR_DESTAQUE)
    _linha_destaque(ws, 6, "SALDO ATUAL", saldo, "C6EFCE" if saldo >= 0 else _COR_DESPESA)

    pct = stats.percentual(despesas_total, renda_total)
    if renda_total:
        ws.cell(row=7, column=1, value=f"Você gastou {pct:.0f}% da sua renda")
    else:
        ws.cell(row=7, column=1, value="Cadastre a renda com /renda 1 <valor> e /renda 2 <valor>")
    ws.cell(row=7, column=1).font = Font(italic=True)

    # Gráfico de rosca: despesas vs. saldo restante
    ws.cell(row=9, column=1, value="Despesas")
    ws.cell(row=9, column=2, value=despesas_total / 100)
    ws.cell(row=10, column=1, value="Saldo")
    ws.cell(row=10, column=2, value=max(saldo, 0) / 100)
    if renda_total:
        grafico = DoughnutChart()
        grafico.title = "Despesas x Saldo"
        dados = Reference(ws, min_col=2, min_row=9, max_row=10)
        rotulos = Reference(ws, min_col=1, min_row=9, max_row=10)
        grafico.add_data(dados)
        grafico.set_categories(rotulos)
        grafico.height, grafico.width = 7, 9
        ws.add_chart(grafico, "D3")

    # Análise de gastos por pessoa
    linha = 13
    ws.cell(row=linha, column=1, value="Análise de Gastos por Pessoa").font = Font(bold=True)
    linha += 1
    for chave, total in stats.total_por_pessoa(gastos_mes):
        ws.cell(row=linha, column=1, value=cfg.nome_pessoa(chave))
        c = ws.cell(row=linha, column=2, value=total / 100)
        c.number_format = _FORMATO_MOEDA
        linha += 1

    # Metas/anotações do mês (para preenchimento manual)
    linha += 1
    ws.cell(row=linha, column=1, value="Metas / anotações do mês 📝").font = Font(bold=True)
    linha += 1
    for _ in range(6):
        ws.cell(row=linha, column=1, value="☐ ")
        linha += 1

    # ================= Gastos Fixos =================
    fixos, variaveis = stats.classificar_fixos_variaveis(gastos_mes, cartoes)
    ws = wb.create_sheet("Gastos Fixos")
    _cabecalho(
        ws,
        ["Responsável", "Categoria", "Descrição", "Data", "Parcela", "Pagamento", "Valor", "Pago?"],
        [16, 14, 26, 12, 10, 16, 14, 8],
    )
    prox = _linhas_gastos(ws, fixos, cfg, 2)
    ws.cell(row=prox + 1, column=6, value="Total:").font = Font(bold=True)
    c = ws.cell(row=prox + 1, column=7, value=stats.total_centavos(fixos) / 100)
    c.number_format, c.font = _FORMATO_MOEDA, Font(bold=True)

    # ================= Gastos do Dia a Dia =================
    ws = wb.create_sheet("Gastos do Dia a Dia")
    _cabecalho(
        ws,
        ["Responsável", "Categoria", "Descrição", "Data", "Parcela", "Pagamento", "Valor", "Pago?"],
        [16, 14, 26, 12, 10, 16, 14, 8],
    )
    prox = _linhas_gastos(ws, variaveis, cfg, 2)
    ws.cell(row=prox + 1, column=6, value="Total:").font = Font(bold=True)
    c = ws.cell(row=prox + 1, column=7, value=stats.total_centavos(variaveis) / 100)
    c.number_format, c.font = _FORMATO_MOEDA, Font(bold=True)

    # ================= Faturas de Cartão do Mês =================
    ws = wb.create_sheet("Faturas de Cartão")
    _cabecalho(ws, ["Cartão", "Total do mês", "Pago?"], [22, 16, 10])
    linha = 2
    cartoes_credito = [c for c in cartoes if c.tipo == "cartao"]
    for c in cartoes_credito:
        gastos_do_cartao = [g for g in gastos_mes if g.cartao_id == c.id]
        total = sum(g.valor_centavos for g in gastos_do_cartao)
        pago = all(g.pago for g in gastos_do_cartao) if gastos_do_cartao else True
        ws.cell(row=linha, column=1, value=c.nome)
        cel = ws.cell(row=linha, column=2, value=total / 100)
        cel.number_format = _FORMATO_MOEDA
        ws.cell(row=linha, column=3, value="Sim" if pago else "Não")
        linha += 1

    # ================= Categorias e Orçamento =================
    ws = wb.create_sheet("Categorias e Orçamento")
    _cabecalho(ws, ["Categoria", "Teto do mês", "Gasto do mês", "% do teto", "Radar"], [18, 16, 16, 12, 8])
    linha = 2
    linha_inicio_grafico = linha
    for r in stats.radar_categorias(gastos_mes, tetos):
        ws.cell(row=linha, column=1, value=r.categoria)
        if r.teto_centavos:
            cel = ws.cell(row=linha, column=2, value=r.teto_centavos / 100)
            cel.number_format = _FORMATO_MOEDA
        cel = ws.cell(row=linha, column=3, value=r.gasto_centavos / 100)
        cel.number_format = _FORMATO_MOEDA
        if r.percentual_do_teto is not None:
            ws.cell(row=linha, column=4, value=round(r.percentual_do_teto) / 100).number_format = "0%"
        ws.cell(row=linha, column=5, value=r.sinal)
        linha += 1
    linha_fim_grafico = linha - 1

    if linha_fim_grafico >= linha_inicio_grafico:
        grafico_cat = BarChart()
        grafico_cat.title = "Planejado x Real"
        grafico_cat.y_axis.title = "R$"
        dados = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=linha_fim_grafico)
        rotulos = Reference(ws, min_col=1, min_row=linha_inicio_grafico, max_row=linha_fim_grafico)
        grafico_cat.add_data(dados, titles_from_data=True)
        grafico_cat.set_categories(rotulos)
        grafico_cat.height, grafico_cat.width = 8, 16
        ws.add_chart(grafico_cat, f"G{linha_inicio_grafico}")

    # ================= Investimentos =================
    ws = wb.create_sheet("Investimentos")
    _cabecalho(ws, ["Local", "Valor"], [24, 16])
    linha = 2
    for local, valor in investimentos:
        ws.cell(row=linha, column=1, value=local)
        ws.cell(row=linha, column=2, value=valor / 100).number_format = _FORMATO_MOEDA
        linha += 1
    ws.cell(row=max(linha, 2) + 1, column=1, value="Total:").font = Font(bold=True)
    c = ws.cell(row=max(linha, 2) + 1, column=2, value=investimentos_total / 100)
    c.number_format, c.font = _FORMATO_MOEDA, Font(bold=True)

    # ================= Cartões e Contas =================
    ws = wb.create_sheet("Cartões e Contas")
    _cabecalho(
        ws,
        ["Nome", "Tipo", "Dia de vencimento", "Dia de fechamento", "Próximo vencimento",
         "Gasto no mês atual", "Último valor registrado"],
        [20, 10, 18, 18, 20, 20, 22],
    )
    por_cartao_mes = dict(stats.total_por_cartao(gastos_mes))
    linha = 2
    for c in cartoes:
        proximo = stats.proximo_vencimento(c, hoje)
        ws.cell(row=linha, column=1, value=c.nome)
        ws.cell(row=linha, column=2, value=c.tipo)
        ws.cell(row=linha, column=3, value=c.dia_vencimento)
        ws.cell(row=linha, column=4, value=c.dia_fechamento)
        if proximo:
            ws.cell(row=linha, column=5, value=proximo).number_format = "DD/MM/YYYY"
        total = por_cartao_mes.get(c.nome, 0)
        ws.cell(row=linha, column=6, value=total / 100).number_format = _FORMATO_MOEDA
        ultimo = db.ultimo_valor_pago(c.id)
        if ultimo is not None:
            ws.cell(row=linha, column=7, value=ultimo / 100).number_format = _FORMATO_MOEDA
        linha += 1

    # ================= Resumo Mensal (histórico) =================
    ws = wb.create_sheet("Resumo Mensal")
    _cabecalho(ws, ["Mês", "Total gasto", "Média mensal (3 meses)"], [14, 16, 22])
    linha = 2
    for inicio, total in stats.serie_mensal(db, hoje, 12):
        ws.cell(row=linha, column=1, value=inicio.strftime("%m/%Y"))
        ws.cell(row=linha, column=2, value=total / 100).number_format = _FORMATO_MOEDA
        linha += 1
    ws.cell(row=2, column=3, value=stats.media_mensal_centavos(db, hoje) / 100).number_format = _FORMATO_MOEDA

    # ================= Resumo Semanal (histórico) =================
    ws = wb.create_sheet("Resumo Semanal")
    _cabecalho(ws, ["Semana (segunda-feira)", "Total gasto", "Média semanal (4 semanas)"], [22, 16, 24])
    linha = 2
    for inicio, total in stats.serie_semanal(db, hoje, 12):
        ws.cell(row=linha, column=1, value=inicio).number_format = "DD/MM/YYYY"
        ws.cell(row=linha, column=2, value=total / 100).number_format = _FORMATO_MOEDA
        linha += 1
    ws.cell(row=2, column=3, value=stats.media_semanal_centavos(db, hoje) / 100).number_format = _FORMATO_MOEDA

    # ================= Lançamentos (histórico completo) =================
    ws = wb.create_sheet("Lançamentos")
    _cabecalho(
        ws,
        ["Data", "Valor", "Estabelecimento", "Categoria", "Cartão/Conta", "Responsável",
         "Parcela", "Pago?", "Descrição", "Origem"],
        [12, 14, 26, 16, 18, 16, 10, 8, 28, 10],
    )
    linha = 2
    for g in db.listar_gastos():
        ws.cell(row=linha, column=1, value=g.data_compra).number_format = "DD/MM/YYYY"
        ws.cell(row=linha, column=2, value=g.valor).number_format = _FORMATO_MOEDA
        ws.cell(row=linha, column=3, value=g.estabelecimento)
        ws.cell(row=linha, column=4, value=g.categoria)
        ws.cell(row=linha, column=5, value=g.cartao_nome or g.forma_pagamento)
        ws.cell(row=linha, column=6, value=cfg.nome_pessoa(g.responsavel) if g.responsavel else "")
        ws.cell(row=linha, column=7, value=g.parcela_rotulo or "")
        ws.cell(row=linha, column=8, value="Sim" if g.pago else "Não")
        ws.cell(row=linha, column=9, value=g.descricao)
        ws.cell(row=linha, column=10, value=g.origem)
        linha += 1

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino
