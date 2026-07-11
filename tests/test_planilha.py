from datetime import date

from openpyxl import load_workbook

from finbot.config import Config
from finbot.db import Database
from finbot.planilha import gerar_planilha


def _config_casal() -> Config:
    return Config(
        telegram_token="x",
        pessoas_por_id={111: "gabriel", 222: "maria_eduarda"},
        nomes_pessoas={"gabriel": "Gabriel", "maria_eduarda": "Maria Eduarda"},
        ordem_pessoas=["gabriel", "maria_eduarda"],
    )


def test_gerar_planilha_completa(tmp_path):
    db = Database(tmp_path / "teste.db")
    cfg = _config_casal()

    nubank = db.adicionar_cartao("Nubank", "cartao", dia_vencimento=10, dia_fechamento=3)
    luz = db.adicionar_cartao("Luz", "conta", dia_vencimento=15)

    db.adicionar_gasto(
        4590,
        date(2026, 7, 8),
        estabelecimento="Supermercado X",
        categoria="mercado",
        cartao_id=nubank.id,
        descricao="compra da semana",
        responsavel="maria_eduarda",
        pago=False,
    )
    db.adicionar_gasto(
        12000,
        date(2026, 7, 9),
        categoria="contas",
        cartao_id=luz.id,
        forma_pagamento="pix",
        responsavel="gabriel",
    )
    db.adicionar_gasto_parcelado(
        120000,
        10,
        date(2026, 7, 1),
        estabelecimento="TV",
        categoria="outros",
        responsavel="combinado",
    )

    db.definir_renda("gabriel", 550000, "Salário")
    db.definir_renda("maria_eduarda", 420000, "Salário")
    db.definir_investimento("Tesouro Selic", 500000)
    db.definir_teto("mercado", 50000)

    destino = tmp_path / "planilha.xlsx"
    gerar_planilha(db, destino, date(2026, 7, 10), cfg)

    wb = load_workbook(destino)
    assert wb.sheetnames == [
        "Resumo do Mês",
        "Gastos Fixos",
        "Gastos do Dia a Dia",
        "Faturas de Cartão",
        "Categorias e Orçamento",
        "Investimentos",
        "Cartões e Contas",
        "Resumo Mensal",
        "Resumo Semanal",
        "Lançamentos",
    ]

    # --- Resumo do Mês: renda, despesas, saldo e "por pessoa" ---
    resumo = wb["Resumo do Mês"]
    total_mes = 4590 + 12000 + 12000  # supermercado + luz + 1ª parcela da TV
    assert resumo.cell(row=3, column=2).value == 970000 / 100  # renda total
    assert resumo.cell(row=4, column=2).value == total_mes / 100
    assert resumo.cell(row=5, column=2).value == 5000.0
    assert resumo.cell(row=6, column=2).value == (970000 - total_mes) / 100

    por_pessoa = {
        resumo.cell(row=r, column=1).value: resumo.cell(row=r, column=2).value
        for r in range(14, 17)
    }
    assert por_pessoa["Gabriel"] == 12000 / 100
    assert por_pessoa["Maria Eduarda"] == 45.90
    assert por_pessoa["Combinado"] == 12000 / 100

    # --- Gastos Fixos: conta cadastrada (Luz) e a parcela da TV ---
    fixos = wb["Gastos Fixos"]
    linhas_fixos = [
        (fixos.cell(row=r, column=1).value, fixos.cell(row=r, column=3).value)
        for r in range(2, fixos.max_row + 1)
        if fixos.cell(row=r, column=1).value
    ]
    assert ("Gabriel", None) in [(p, d) for p, d in linhas_fixos if p == "Gabriel"]
    assert any(p == "Combinado" and d == "TV" for p, d in linhas_fixos)
    # a parcela aparece com o rótulo "1/10"
    parcela_tv = [
        fixos.cell(row=r, column=5).value
        for r in range(2, fixos.max_row + 1)
        if fixos.cell(row=r, column=3).value == "TV"
    ]
    assert parcela_tv == ["1/10"]

    # --- Gastos do Dia a Dia: supermercado (Nubank, não é "conta") ---
    variaveis = wb["Gastos do Dia a Dia"]
    assert variaveis.cell(row=2, column=1).value == "Maria Eduarda"
    assert variaveis.cell(row=2, column=3).value == "Supermercado X"
    assert variaveis.cell(row=2, column=8).value == "Não"  # pago=False

    # --- Faturas de Cartão: só cartões tipo 'cartao' (Nubank) ---
    faturas = wb["Faturas de Cartão"]
    nomes_fatura = [faturas.cell(row=r, column=1).value for r in range(2, faturas.max_row + 1)]
    assert nomes_fatura == ["Nubank"]
    assert faturas.cell(row=2, column=2).value == 45.90
    assert faturas.cell(row=2, column=3).value == "Não"

    # --- Categorias e Orçamento: teto de mercado configurado ---
    categorias = wb["Categorias e Orçamento"]
    linha_mercado = next(
        r for r in range(2, categorias.max_row + 1)
        if categorias.cell(row=r, column=1).value == "mercado"
    )
    assert categorias.cell(row=linha_mercado, column=2).value == 500.0
    assert categorias.cell(row=linha_mercado, column=3).value == 45.90

    # --- Investimentos ---
    investimentos = wb["Investimentos"]
    assert investimentos.cell(row=2, column=1).value == "Tesouro Selic"
    assert investimentos.cell(row=2, column=2).value == 5000.0

    # --- Cartões e Contas (mantido do formato anterior) ---
    cartoes = wb["Cartões e Contas"]
    nomes = {cartoes.cell(row=r, column=1).value for r in (2, 3)}
    assert nomes == {"Nubank", "Luz"}

    # --- Lançamentos: histórico completo com as novas colunas ---
    lanc = wb["Lançamentos"]
    assert lanc.cell(row=1, column=6).value == "Responsável"
    responsaveis = {lanc.cell(row=r, column=6).value for r in range(2, lanc.max_row + 1)}
    assert responsaveis == {"Gabriel", "Maria Eduarda", "Combinado"}


def test_gerar_planilha_banco_vazio(tmp_path):
    """Não deve quebrar quando não há gastos, renda ou pessoas configuradas."""
    db = Database(tmp_path / "vazio.db")
    cfg = Config(telegram_token="x")

    destino = tmp_path / "vazia.xlsx"
    gerar_planilha(db, destino, date(2026, 7, 10), cfg)

    wb = load_workbook(destino)
    resumo = wb["Resumo do Mês"]
    assert resumo.cell(row=3, column=2).value == 0.0
    assert resumo.cell(row=6, column=2).value == 0.0
